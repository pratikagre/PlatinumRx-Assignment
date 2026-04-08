import os
from openpyxl import Workbook

# Ensure the Spreadsheets directory exists
os.makedirs('../Spreadsheets', exist_ok=True)

wb = Workbook()

# Create ticket sheet
ws1 = wb.active
ws1.title = "ticket"

ws1.append(["ticket_id", "created_at", "closed_at", "outlet_id", "cms_id"])
ws1.append(["isu-sjd-457", "2021-08-19 16:45:43", "2021-08-22 12:33:32", "wrqy-juv-978", "vew-iuvd-12"])
ws1.append(["qer-fal-092", "2021-08-21 11:09:22", "2021-08-21 17:13:45", "8woh-k3u-23b", "cms-2384-xyz"])
ws1.append(["tk-394-xfe", "2021-08-21 13:20:00", "2021-08-21 13:45:00", "wrqy-juv-978", "vew-iuvd-15"])
ws1.append(["tk-395-ytr", "2021-08-22 10:00:00", "2021-08-22 10:30:00", "8woh-k3u-23b", "vew-iuvd-16"])
ws1.append(["tk-396-ewq", "2021-08-22 14:00:00", "2021-08-23 09:00:00", "wrqy-juv-978", "vew-iuvd-17"])

# Create feedbacks sheet
ws2 = wb.create_sheet(title="feedbacks")
ws2.append(["cms_id", "feedback_at", "feedback_rating", "ticket_created_at"])
ws2.append(["vew-iuvd-12", "2021-08-21 13:26:48", 3, '=INDEX(ticket!B:B, MATCH(A2, ticket!E:E, 0))'])
ws2.append(["cms-2384-xyz", "2021-08-22 09:12:00", 5, '=INDEX(ticket!B:B, MATCH(A3, ticket!E:E, 0))'])
ws2.append(["vew-iuvd-15", "2021-08-23 10:15:00", 4, '=INDEX(ticket!B:B, MATCH(A4, ticket!E:E, 0))'])
ws2.append(["vew-iuvd-16", "2021-08-24 11:20:00", 2, '=INDEX(ticket!B:B, MATCH(A5, ticket!E:E, 0))'])

# Add helper columns for ticket to show same day and same hour count logic
ws1["F1"] = "Same Day?"
ws1["F2"] = '=IF(INT(DATEVALUE(LEFT(B2,10)))=INT(DATEVALUE(LEFT(C2,10))), TRUE, FALSE)'
ws1["F3"] = '=IF(INT(DATEVALUE(LEFT(B3,10)))=INT(DATEVALUE(LEFT(C3,10))), TRUE, FALSE)'
ws1["F4"] = '=IF(INT(DATEVALUE(LEFT(B4,10)))=INT(DATEVALUE(LEFT(C4,10))), TRUE, FALSE)'
ws1["F5"] = '=IF(INT(DATEVALUE(LEFT(B5,10)))=INT(DATEVALUE(LEFT(C5,10))), TRUE, FALSE)'
ws1["F6"] = '=IF(INT(DATEVALUE(LEFT(B6,10)))=INT(DATEVALUE(LEFT(C6,10))), TRUE, FALSE)'

ws1["G1"] = "Same Hour?"
ws1["G2"] = '=IF(AND(F2=TRUE, HOUR(TIMEVALUE(MID(B2, 12, 8)))=HOUR(TIMEVALUE(MID(C2, 12, 8)))), TRUE, FALSE)'
ws1["G3"] = '=IF(AND(F3=TRUE, HOUR(TIMEVALUE(MID(B3, 12, 8)))=HOUR(TIMEVALUE(MID(C3, 12, 8)))), TRUE, FALSE)'
ws1["G4"] = '=IF(AND(F4=TRUE, HOUR(TIMEVALUE(MID(B4, 12, 8)))=HOUR(TIMEVALUE(MID(C4, 12, 8)))), TRUE, FALSE)'
ws1["G5"] = '=IF(AND(F5=TRUE, HOUR(TIMEVALUE(MID(B5, 12, 8)))=HOUR(TIMEVALUE(MID(C5, 12, 8)))), TRUE, FALSE)'
ws1["G6"] = '=IF(AND(F6=TRUE, HOUR(TIMEVALUE(MID(B6, 12, 8)))=HOUR(TIMEVALUE(MID(C6, 12, 8)))), TRUE, FALSE)'

# Auto-adjust column widths and set header bold for ticket sheet
from openpyxl.styles import Font
bold_font = Font(bold=True)
for cell in ws1[1]:
    cell.font = bold_font

column_widths = {'A': 15, 'B': 22, 'C': 22, 'D': 15, 'E': 15, 'F': 12, 'G': 12}
for col, width in column_widths.items():
    ws1.column_dimensions[col].width = width

# Same formatting for feedbacks sheet
for cell in ws2[1]:
    cell.font = bold_font
ws2.column_dimensions['A'].width = 15
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 18
ws2.column_dimensions['D'].width = 30

# Create Pivot Table Summary Sheet via COUNTIFS logic
ws3 = wb.create_sheet(title="Pivot Summary")
ws3.append(["outlet_id", "Same Day Ticket Count", "Same Hour Ticket Count"])
ws3.append(["wrqy-juv-978", '=COUNTIFS(ticket!D:D, A2, ticket!F:F, TRUE)', '=COUNTIFS(ticket!D:D, A2, ticket!G:G, TRUE)'])
ws3.append(["8woh-k3u-23b", '=COUNTIFS(ticket!D:D, A3, ticket!F:F, TRUE)', '=COUNTIFS(ticket!D:D, A3, ticket!G:G, TRUE)'])

for cell in ws3[1]:
    cell.font = bold_font
ws3.column_dimensions['A'].width = 20
ws3.column_dimensions['B'].width = 25
ws3.column_dimensions['C'].width = 25
# Save Excel
wb.save(r'../Spreadsheets/Ticket_Analysis.xlsx')
print("Successfully created Ticket_Analysis.xlsx")
